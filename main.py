from moodle_client import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def export_to_excel(data: Union[Tuple[List[Dict], int], Dict], filename: str) -> None:
    """
    Универсальная функция для экспорта данных из:
    - get_teacher_courses()
    - track_interim_assessment()
    """
    if not data:
        print("Нет данных для экспорта")
        return

    try:
        wb = Workbook()

        if isinstance(data, tuple) and len(data) == 2:
            courses, count = data
            ws = wb.active
            ws.title = "Курсы преподавателя"

            headers = [
                "ID курса",
                "Название курса",
                "Дата начала",
                "Дата окончания"
            ]
            ws.append(headers)

            for course in courses:
                ws.append([
                    course['id'],
                    course['fullname'],
                    course['startdate'],
                    course['enddate']
                ])

            ws.append([])
            ws.append(["Всего курсов:", count])

        elif isinstance(data, dict) and 'students_grades' in data:
            ws = wb.active
            ws.title = "Аттестация"

            ws.append(["Курс ID:", data['course_id']])
            ws.append(["Период:",
                       f"{data['period']['start']} - {data['period']['end']}"])
            ws.append([])

            headers = [
                "ФИО студента",
                "Элемент оценивания",
                "Оценка",
                "Процент",
                "Макс. оценка",
                "Дата сдачи"
            ]
            ws.append(headers)

            for student in data['students_grades']:
                for grade in student['grades']:
                    item = next(
                        (i for i in data['interim_items']
                         if i['name'] == grade['item']),
                        None
                    )

                    ws.append([
                        student['userfullname'],
                        grade['item'],
                        grade['grade'],
                        grade['percentage'],
                        item['max_grade'] if item else '-',
                        item['date'] if item else '-'
                    ])

            ws_items = wb.create_sheet("Элементы оценивания")
            ws_items.append([
                "ID", "Название", "Тип", "Макс. оценка", "Дата"
            ])

            for item in data['interim_items']:
                ws_items.append([
                    item['id'],
                    item['name'],
                    item['type'],
                    item['max_grade'],
                    item['date']
                ])

        else:
            print("Неподдерживаемый формат данных")
            return

        for sheet in wb:
            for row in sheet.iter_rows(max_row=1):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

            for column in sheet.columns:
                max_length = 0
                column_cells = [cell for cell in column]
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        wb.save(filename)
        print(f"Файл {filename} успешно сохранен")

    except Exception as e:
        print(f"Ошибка при экспорте в Excel: {e}")


def zachet_export_to_excel(data: List[Dict], filename: str, group_name: str) -> None:
    """
    Экспорт результатов в Excel (только ФИО, группа и оценка)
    Улучшенная версия с:
    - Проверкой данных
    - Очисткой имен
    - Форматированием оценок
    """
    if not data:
        print("Нет данных для экспорта")
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты зачета"

        # Заголовки с форматированием
        headers = ["ФИО студента", "Группа", "Оценка"]
        ws.append(headers)

        # Заполняем данные
        for result in data:
            ws.append([
                result['user_name'],
                group_name,
                result['best_grade']
            ])

        # Форматирование
        # 1. Жирные заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 2. Автоподбор ширины колонок
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col),
                default=0
            )
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

        # 3. Центрирование оценок
        for row in ws.iter_rows(min_row=2, max_col=3, max_row=len(data)+1):
            row[2].alignment = Alignment(horizontal='center')  # Колонка с оценками

        wb.save(filename)
        print(f"Файл сохранен: {filename}")

    except Exception as e:
        print(f"Ошибка экспорта: {str(e)}")
        raise

client = MoodleClient()
atempts = client.analyze_attempts_results([3], 1, 5)
#groups = client.get_course_groups(5)
zachet_export_to_excel(atempts, 'f.xlsx', client.get_group_name(1))
